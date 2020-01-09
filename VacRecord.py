from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import worksheet
import datetime, time


def VacationsConsumedList():
  wb = load_workbook('Hazأجازات2020.xlsx', data_only=True)
  ws = wb['Sheet1']
  VacationsConsumed=[]
  for i in ws["F3:F12"]:
     VacationsConsumed.append(i[0].value)

  print(VacationsConsumed)
  return(VacationsConsumed)

def GetDataEmployee(ListOfEmployees):
    wb = load_workbook('1_StandardReport.xlsx', data_only=True)
    ws = wb['Att.log report']
    k=0
    for i in range (8,27,2):
        for j in range(1,7):
            ListOfEmployees[k].append(ws.cell(row=i, column= j).value)
        k=k+1
    print(ListOfEmployees)

def GetDateReport():
    wb = load_workbook('1_StandardReport.xlsx', data_only=True)
    ws = wb['Att.log report']
    DateReport=datetime.datetime.strptime(ws.cell(row=3,column=3).value[0:10],"%Y-%m-%d")
    #print(DateReport.day)
    return DateReport

def NumberOfEmployees():
    return len(ListOfEmployees)



#GetDateRecord function is responsible for determining the location of the first cell under the first day in the week
#in the vacations record
def GetDateRecord():
    wb = load_workbook('Hazأجازات2020.xlsx', data_only=True)
    ws = wb['Sheet1']
    # date in yyyy/mm/dd format
    #d1 = datetime.datetime(2020, 1, 1)
    d1=GetDateReport()
    #Look for the cell that has the starting date

    for j in range(1,373):
        for i in range(1,3):
            if ws.cell(row=i , column=j).value == d1:
               x =i+1
               y =j
            else:
                continue
    return x,y
#FillDateRecord is responsible for going through the data of the ListOfEmployees and get out all the "None" values and
#filling them as v in the record
def FillDateRecord():
    wb = load_workbook('Hazأجازات20202.xlsx', data_only=True)
    ws = wb['Sheet1']
    x=GetDateRecord()[0]
    y=GetDateRecord()[1]
    for i in range(NumberOfEmployees()):
        for j in range(6):
            if ListOfEmployees[i][j] is not None:
                ws.cell(row=x, column=y).value = ''
                y=y+1
            else:
                ws.cell(row=x, column=y).value = 'v'
                y=y+1
        x=x+1
        y=GetDateRecord()[1]
    print(x,y)

    wb.save('Haz20202أجازات.xlsx')

def fillin_functions():
    wb = load_workbook('Hazأجازات20202.xlsx', data_only=True)
    ws = wb['Sheet1']
    


Anwar, Amr, Hussein, Taha, Hashim, Lotfy, Aly, Saad, Tarek, Mostafa =[],[],[],[],[],[],[],[],[],[]
ListOfEmployees=[      Anwar, Amr, Hussein, Taha, Hashim, Lotfy,   Aly,   Saad, Tarek,   Mostafa]

print(GetDateRecord())
GetDataEmployee(ListOfEmployees)
FillDateRecord()



